from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .localization import display_label, display_value


THIN = Side(style="thin", color="A9B4C2")
MEDIUM = Side(style="medium", color="457B9D")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

BP_BLUE = "457B9D"
BP_CYAN = "A8DADC"
BP_PALE = "F1FAEE"
BP_PRO = "B8C3C8"
PLAYER_BLUE = "4A90E2"
CO_TITLE = "D1F2EB"
CO_HEADER = "A9DFBF"
CO_NOTE = "FDFEFE"
WARD_TITLE = "D6EAF8"
WARD_NOTE = "EBF5FB"
WARD_HEADER = "3498DB"
WIN_GREEN = "27AE60"
LOSS_RED = "E74C3C"


def _sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name)).strip() or "数据"
    clean = clean[:31]
    candidate = clean
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{clean[:31-len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _cell_value(value: Any, column: str = "") -> Any:
    normalized = column.lower().replace(" ", "_")
    if value is not None and (
        "account_id" in normalized
        or "match_id" in normalized
        or column in {"Account ID", "Match ID", "Steam ID"}
    ):
        return str(value)
    if value is None or isinstance(value, (str, int, float, bool, datetime)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _columns(rows: Sequence[Dict[str, Any]]) -> List[str]:
    result: List[str] = []
    for row in rows:
        for key in row:
            if key not in result:
                result.append(key)
    return result


def _apply_number_format(cell: Any, column: str, value: Any) -> None:
    if value is None:
        return
    if "胜率(%)" in column or "被反率(%)" in column or "占比(%)" in column:
        cell.number_format = '0.0"%"'
    elif any(word in column for word in ("平均", "KDA", "每分钟", "时长(分)")):
        cell.number_format = "0.00"
    elif isinstance(value, int):
        cell.number_format = "#,##0"


def _set_widths(
    worksheet: Any,
    columns: Sequence[str],
    start_col: int = 1,
    rows: Sequence[Dict[str, Any]] | None = None,
) -> None:
    materialized = list(rows or [])
    for offset, column in enumerate(columns):
        values = [str(row.get(column) or "") for row in materialized[:250]]
        maximum = max([len(str(column)), *(len(value) for value in values)], default=10)
        if "阵容" in column or "组合" in column:
            width = min(max(maximum + 2, 30), 50)
        elif column in {"内容", "对手队伍"}:
            width = min(max(maximum + 2, 20), 48)
        else:
            width = min(max(maximum + 2, 11), 28)
        worksheet.column_dimensions[
            get_column_letter(start_col + offset)
        ].width = width


def _write_table(
    worksheet: Any,
    rows: Iterable[Dict[str, Any]],
    *,
    start_row: int,
    start_col: int = 1,
    header_fill: str,
    header_font_color: str = "FFFFFF",
    data_fill: str | None = None,
    title: str | None = None,
    title_fill: str | None = None,
    note: str | None = None,
    note_fill: str | None = None,
    note_font_color: str = "1B4F72",
    filter_table: bool = True,
) -> tuple[int, List[str]]:
    materialized = list(rows)
    columns = _columns(materialized)
    if not columns:
        columns = ["说明"]
        materialized = [{"说明": "暂无数据"}]
    title_rows = 0
    if title:
        end_col = start_col + len(columns) - 1
        worksheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=end_col,
        )
        cell = worksheet.cell(start_row, start_col, title)
        cell.fill = PatternFill("solid", fgColor=title_fill or header_fill)
        cell.font = Font(bold=True, size=14)
        cell.alignment = CENTER
        cell.border = ALL_BORDER
        worksheet.row_dimensions[start_row].height = 26
        title_rows += 1
    if note:
        row = start_row + title_rows
        end_col = start_col + len(columns) - 1
        worksheet.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col,
        )
        cell = worksheet.cell(row, start_col, note)
        cell.fill = PatternFill("solid", fgColor=note_fill or "FFFFFF")
        cell.font = Font(italic=True, color=note_font_color)
        cell.alignment = CENTER
        cell.border = ALL_BORDER
        worksheet.row_dimensions[row].height = 32
        title_rows += 1
    header_row = start_row + title_rows
    for offset, column in enumerate(columns):
        cell = worksheet.cell(
            header_row, start_col + offset, display_label(column)
        )
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=True, color=header_font_color)
        cell.alignment = CENTER
        cell.border = ALL_BORDER
    worksheet.row_dimensions[header_row].height = 30
    for row_offset, row in enumerate(materialized, start=1):
        excel_row = header_row + row_offset
        for col_offset, column in enumerate(columns):
            value = _cell_value(display_value(row.get(column)), column)
            cell = worksheet.cell(excel_row, start_col + col_offset, value)
            cell.border = ALL_BORDER
            cell.alignment = CENTER
            if data_fill:
                cell.fill = PatternFill("solid", fgColor=data_fill)
            _apply_number_format(cell, column, value)
    end_row = header_row + len(materialized)
    if filter_table:
        worksheet.auto_filter.ref = (
            f"{get_column_letter(start_col)}{header_row}:"
            f"{get_column_letter(start_col + len(columns) - 1)}{end_row}"
        )
    _set_widths(worksheet, columns, start_col, materialized)
    return end_row, columns


def _write_raw_sheet(
    workbook: Workbook,
    used_names: set[str],
    name: str,
    rows: Iterable[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(_sheet_name(name, used_names))
    _write_table(
        worksheet,
        rows,
        start_row=1,
        header_fill="16324F",
        title=name,
        title_fill="D7E3F0",
    )
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False


def export_hero_meta_workbook(
    rows: Sequence[Dict[str, Any]],
    path: str | Path,
    *,
    title: str = "Dota2 版本英雄胜率数据",
    note: str = "",
) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "版本英雄数据"
    worksheet.sheet_view.showGridLines = False
    end_row, columns = _write_table(
        worksheet,
        list(rows),
        start_row=1,
        header_fill="0969DA",
        title=title,
        title_fill="DDF4FF",
        note=note
        or "说明：职业比赛与玩家公开对局为不同口径；小版本数据基于 OpenDota match_patch 聚合。",
        note_fill="F6F8FA",
        note_font_color="57606A",
    )
    worksheet.freeze_panes = "A4"
    if rows and "英雄名称" in columns and any(
        key in columns for key in ("场次", "选取次数")
    ):
        value_name = "场次" if "场次" in columns else "选取次数"
        category_col = columns.index("英雄名称") + 1
        value_col = columns.index(value_name) + 1
        last_data_row = min(end_row, 3 + min(len(rows), 15))
        _add_bar_chart(
            worksheet,
            title=f"前15英雄{value_name}",
            header_row=3,
            first_data_row=4,
            last_data_row=last_data_row,
            category_col=category_col,
            value_cols=(value_col,),
            anchor=f"{get_column_letter(len(columns) + 2)}4",
        )
    workbook.save(output)
    return output


def _add_bar_chart(
    worksheet: Any,
    *,
    title: str,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
    category_col: int,
    value_cols: Sequence[int],
    anchor: str,
) -> None:
    if last_data_row < first_data_row:
        return
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "次数"
    chart.x_axis.title = "英雄"
    for column in value_cols:
        data = Reference(
            worksheet,
            min_col=column,
            min_row=header_row,
            max_row=last_data_row,
        )
        chart.add_data(data, titles_from_data=True)
    categories = Reference(
        worksheet,
        min_col=category_col,
        min_row=first_data_row,
        max_row=last_data_row,
    )
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 18
    worksheet.add_chart(chart, anchor)


def _write_bp_module(
    workbook: Workbook,
    used_names: set[str],
    dataset: Dict[str, Any],
    module: Dict[str, Any],
) -> None:
    team_name = dataset.get("team", {}).get("name") or "目标战队"
    stats = dataset.get("statistics", {})
    summary_rows = [
        {"指标": "战队名称", "值": team_name},
        {"指标": "请求分析场次", "值": stats.get("matches_requested")},
        {"指标": "实际分析场次", "值": stats.get("matches_analyzed")},
        {"指标": "总胜场", "值": stats.get("wins")},
        {"指标": "总胜率", "值": f"{stats.get('win_rate', 0):.2f}%"},
        {
            "指标": "有效 BP 比赛",
            "值": module.get("quality", {}).get("有效BP比赛", 0),
        },
    ]
    worksheet = workbook.create_sheet(_sheet_name("0. BP报告摘要", used_names))
    end_row, _ = _write_table(
        worksheet,
        summary_rows,
        start_row=1,
        header_fill=BP_BLUE,
        title="BP 分析报告摘要",
        title_fill=BP_CYAN,
        filter_table=False,
    )
    note_row = end_row + 2
    worksheet.cell(note_row, 1, "--- 数据来源说明 ---")
    worksheet.cell(note_row, 1).font = Font(bold=True)
    worksheet.cell(note_row + 1, 1, "比赛与 BP 数据均由 OpenDota API 实时抓取。")
    worksheet.cell(
        note_row + 2,
        1,
        "若公开比赛详情缺少 picks_bans，则该场不会进入 BP 统计。",
    )
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 28
    worksheet.sheet_view.showGridLines = False

    summary = list(module.get("summary", []))
    bans = [
        {
            "英雄名称": row["英雄"],
            "禁用总次数": row.get("我方禁用", 0),
        }
        for row in summary
        if row.get("我方禁用", 0)
    ][:20]
    picks = [
        {
            "英雄名称": row["英雄"],
            "选取总次数": row.get("我方选取", 0),
            "选取后胜率": row.get("选取胜率(%)"),
        }
        for row in summary
        if row.get("我方选取", 0)
    ][:20]
    worksheet = workbook.create_sheet(_sheet_name("1. 所有BP英雄概览", used_names))
    left_end, left_columns = _write_table(
        worksheet,
        bans,
        start_row=1,
        start_col=1,
        header_fill=BP_BLUE,
        title="禁用英雄总览 - 前 20",
        title_fill=BP_CYAN,
        filter_table=False,
    )
    right_start = len(left_columns) + 2
    _write_table(
        worksheet,
        picks,
        start_row=1,
        start_col=right_start,
        header_fill=BP_BLUE,
        title="选择英雄总览 - 前 20",
        title_fill=BP_CYAN,
        filter_table=False,
    )
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False

    opponent_bans = [
        {
            "英雄名称": row["英雄"],
            "对手禁用次数": row.get("对手禁用", 0),
            "我方选取次数": row.get("我方选取", 0),
            "我方选取胜率(%)": row.get("选取胜率(%)"),
        }
        for row in summary
        if row.get("对手禁用", 0)
    ][:20]
    worksheet = workbook.create_sheet(_sheet_name("2. 对手禁用统计", used_names))
    end_row, _ = _write_table(
        worksheet,
        opponent_bans,
        start_row=1,
        header_fill=BP_BLUE,
        title="对手禁用英雄统计（前 20）",
        title_fill=BP_CYAN,
    )
    worksheet.freeze_panes = "A4"
    _add_bar_chart(
        worksheet,
        title="对手禁用次数 Top 20",
        header_row=2,
        first_data_row=3,
        last_data_row=end_row,
        category_col=1,
        value_cols=(2,),
        anchor=f"B{end_row + 3}",
    )
    worksheet.sheet_view.showGridLines = False

    first_last = [
        {
            "英雄名称": row["英雄"],
            "首轮禁用次数": row.get("首轮禁用", 0),
            "首轮选择次数": row.get("首轮选取", 0),
            "末轮禁用次数": row.get("末轮禁用", 0),
            "末轮选择次数": row.get("末轮选取", 0),
        }
        for row in summary
        if any(
            row.get(key, 0)
            for key in ("首轮禁用", "首轮选取", "末轮禁用", "末轮选取")
        )
    ][:15]
    worksheet = workbook.create_sheet(_sheet_name("3. 首末轮BP统计", used_names))
    end_row, _ = _write_table(
        worksheet,
        first_last,
        start_row=1,
        header_fill=BP_BLUE,
        title="目标队伍首末轮 BP 统计（前 15）",
        title_fill=BP_CYAN,
    )
    worksheet.freeze_panes = "A4"
    _add_bar_chart(
        worksheet,
        title="首轮与末轮 BP 比较",
        header_row=2,
        first_data_row=3,
        last_data_row=end_row,
        category_col=1,
        value_cols=(2, 3, 4, 5),
        anchor=f"B{end_row + 3}",
    )
    worksheet.sheet_view.showGridLines = False

    timeline = list(module.get("lost_timeline", []))
    worksheet = workbook.create_sheet(_sheet_name("4. 输掉比赛BP时间轴", used_names))
    end_row, columns = _write_table(
        worksheet,
        timeline,
        start_row=1,
        header_fill=BP_BLUE,
        title="输掉比赛的 BP 时间轴详情",
        title_fill=BP_CYAN,
    )
    previous_match = None
    match_col = columns.index("Match ID") + 1 if "Match ID" in columns else None
    actor_col = columns.index("执行方") + 1 if "执行方" in columns else None
    for row_number in range(3, end_row + 1):
        match_id = worksheet.cell(row_number, match_col).value if match_col else None
        actor = worksheet.cell(row_number, actor_col).value if actor_col else None
        if previous_match is not None and match_id != previous_match:
            for cell in worksheet[row_number]:
                cell.border = Border(top=MEDIUM, left=THIN, right=THIN, bottom=THIN)
        fill = BP_CYAN if actor == "我方" else BP_PALE
        for cell in worksheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=fill)
        previous_match = match_id
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False

    _write_raw_sheet(
        workbook,
        used_names,
        "BP-全部动作原始数据",
        module.get("actions", []),
    )


def _write_player_module(
    workbook: Workbook,
    used_names: set[str],
    module: Dict[str, Any],
) -> None:
    overview_by_id = {
        int(row["Account ID"]): row for row in module.get("overview", [])
    }
    profiles = list(module.get("profiles", []))
    recent_heroes = list(module.get("recent_heroes", []))
    career_heroes = list(module.get("career_heroes", []))
    for profile in profiles:
        account_id = int(profile["Account ID"])
        recent = overview_by_id.get(account_id, {})
        display_name = (
            profile.get("职业名称")
            or profile.get("昵称")
            or recent.get("选手")
            or f"ID_{account_id}"
        )
        worksheet = workbook.create_sheet(_sheet_name(display_name, used_names))
        worksheet.sheet_view.showGridLines = False
        worksheet.cell(1, 1, f"OpenDota 个人档案链接 (队员: {display_name}):")
        link = worksheet.cell(1, 2, "点击查看档案")
        link.hyperlink = f"https://www.opendota.com/players/{account_id}"
        link.style = "Hyperlink"
        current_row = 3

        profile_rows = [{"字段": key, "值": value} for key, value in profile.items()]
        current_row, _ = _write_table(
            worksheet,
            profile_rows,
            start_row=current_row,
            header_fill=PLAYER_BLUE,
            title="--- 表格 1: OpenDota 生涯档案概观 (Profile & Overview) ---",
            title_fill="FFFFFF",
            filter_table=False,
        )
        current_row += 2
        recent_rows = [recent] if recent else []
        current_row, _ = _write_table(
            worksheet,
            recent_rows,
            start_row=current_row,
            header_fill=PLAYER_BLUE,
            title="--- 表格 2: 职业队内近期综合表现 ---",
            title_fill="FFFFFF",
            filter_table=False,
        )
        current_row += 2
        player_recent_heroes = [
            row
            for row in recent_heroes
            if int(row.get("Account ID") or 0) == account_id
        ]
        recent_start = current_row
        current_row, recent_columns = _write_table(
            worksheet,
            player_recent_heroes,
            start_row=current_row,
            header_fill=PLAYER_BLUE,
            title="--- 表格 3: 职业队内近期英雄表现 ---",
            title_fill="FFFFFF",
            filter_table=False,
        )
        if player_recent_heroes and "英雄" in recent_columns:
            header_row = recent_start + 1
            first_data = header_row + 1
            category_col = recent_columns.index("英雄") + 1
            value_col = recent_columns.index("场次") + 1
            _add_bar_chart(
                worksheet,
                title="近期英雄使用次数",
                header_row=header_row,
                first_data_row=first_data,
                last_data_row=current_row,
                category_col=category_col,
                value_cols=(value_col,),
                anchor=f"G{recent_start}",
            )
        current_row += 2
        player_career_heroes = [
            row
            for row in career_heroes
            if int(row.get("Account ID") or 0) == account_id
        ][:20]
        _write_table(
            worksheet,
            player_career_heroes,
            start_row=current_row,
            header_fill=PLAYER_BLUE,
            title="--- 表格 4: OpenDota 生涯英雄统计 (生涯前20英雄) ---",
            title_fill="FFFFFF",
            filter_table=False,
        )
        worksheet.freeze_panes = "A3"
        worksheet.column_dimensions["A"].width = max(
            worksheet.column_dimensions["A"].width or 0, 30
        )

    if not profiles:
        _write_raw_sheet(
            workbook, used_names, "选手-近期概览", module.get("overview", [])
        )


def _write_cooccurrence_module(
    workbook: Workbook,
    used_names: set[str],
    module: Dict[str, Any],
) -> None:
    definitions = (
        (
            "多英雄共现自由检索",
            module.get("match_lineups", []),
            "多英雄共现自由检索",
            "说明：逐场查看我方阵容、对手阵容、对手队伍与比赛结果。",
        ),
        (
            "阵容共现与胜率分析",
            module.get("combinations", []),
            "阵容共现与胜率分析 (2-5人)",
            "说明：聚合统计 2-5 人英雄组合的出现次数、胜场与胜率。",
        ),
        (
            "英雄共现热度",
            module.get("hero_heat", []),
            "英雄共现热度",
            "说明：统计英雄上场频率、占比和共现加权度。",
        ),
    )
    for name, rows, title, note in definitions:
        worksheet = workbook.create_sheet(_sheet_name(name, used_names))
        _write_table(
            worksheet,
            rows,
            start_row=1,
            header_fill=CO_HEADER,
            header_font_color="000000",
            title=title,
            title_fill=CO_TITLE,
            note=note,
            note_fill=CO_NOTE,
            note_font_color="117A65",
        )
        worksheet.freeze_panes = "A4"
        worksheet.sheet_view.showGridLines = False


def _write_ward_sheet(
    workbook: Workbook,
    used_names: set[str],
    *,
    name: str,
    rows: Iterable[Dict[str, Any]],
    title: str,
    note: str,
    divider: bool = False,
) -> None:
    materialized = list(rows)
    worksheet = workbook.create_sheet(_sheet_name(name, used_names))
    end_row, columns = _write_table(
        worksheet,
        materialized,
        start_row=1,
        header_fill=WARD_HEADER,
        title=title,
        title_fill=WARD_TITLE,
        note=note,
        note_fill=WARD_NOTE,
    )
    result_col = columns.index("结果") + 1 if "结果" in columns else None
    match_col = columns.index("Match ID") + 1 if "Match ID" in columns else None
    previous_match = None
    for row_number in range(4, end_row + 1):
        result = worksheet.cell(row_number, result_col).value if result_col else None
        match_id = worksheet.cell(row_number, match_col).value if match_col else None
        if result in {"胜", "WIN"}:
            cell = worksheet.cell(row_number, result_col)
            cell.fill = PatternFill("solid", fgColor=WIN_GREEN)
            cell.font = Font(color="FFFFFF", bold=True)
        elif result in {"负", "LOSS"}:
            cell = worksheet.cell(row_number, result_col)
            cell.fill = PatternFill("solid", fgColor=LOSS_RED)
            cell.font = Font(color="FFFFFF", bold=True)
        if divider and previous_match is not None and match_id != previous_match:
            for cell in worksheet[row_number]:
                cell.border = Border(top=MEDIUM, left=THIN, right=THIN, bottom=THIN)
        previous_match = match_id
    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False


def _write_ward_module(
    workbook: Workbook,
    used_names: set[str],
    team_name: str,
    module: Dict[str, Any],
) -> None:
    _write_ward_sheet(
        workbook,
        used_names,
        name="1-概览",
        rows=module.get("overview", []),
        title=f"{team_name} · 眼位概览（v1.6b 口径）",
        note=(
            "说明：每场统计双方假/真眼数量。判定规则：假眼≥355秒、"
            "真眼≥415秒视为自然消失；无 left_event 时按原脚本容错。"
        ),
    )
    _write_ward_sheet(
        workbook,
        used_names,
        name="2-我方逐场明细",
        rows=module.get("mine_details", []),
        title="我方逐场眼位明细（含消失时间/持续时间/消失类型）",
        note="说明：可筛选 Match ID、阶段、类型、玩家和消失类型。",
        divider=True,
    )
    _write_ward_sheet(
        workbook,
        used_names,
        name="3-对手逐场明细",
        rows=module.get("opponent_details", []),
        title="对手逐场眼位明细（含消失时间/持续时间/消失类型）",
        note="说明：用于对比双方视野质量与拆眼效率。",
        divider=True,
    )
    _write_ward_sheet(
        workbook,
        used_names,
        name="4-热力格点矩阵",
        rows=module.get("heat_grid", []),
        title="热力格点矩阵（64×64）",
        note="说明：逐场格点计数，按我方/对手 × 假眼/真眼区分。",
        divider=True,
    )
    _write_ward_sheet(
        workbook,
        used_names,
        name="5-存活质量统计",
        rows=module.get("quality", []),
        title="存活质量统计（每分钟插/拆眼；分段被反率与存活时长）",
        note="说明：拆眼/分钟以对手眼位被提前移除数量近似；分段质量指标针对假眼。",
        divider=True,
    )


def _quality_rows(dataset: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = [
        {"项目": key, "内容": _cell_value(value)}
        for key, value in dataset.get("data_quality", {}).items()
    ]
    for module_name, module in dataset.get("modules", {}).items():
        for key in ("quality", "data_quality"):
            if isinstance(module.get(key), dict):
                rows.extend(
                    {
                        "项目": f"{module_name}.{item_key}",
                        "内容": _cell_value(item_value),
                    }
                    for item_key, item_value in module[key].items()
                )
    return rows


def export_dataset_to_excel(dataset: Dict[str, Any], output_path: str) -> Path:
    target = Path(output_path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()

    if dataset.get("sheet_count") is not None:
        for sheet in dataset.get("sheets", []):
            _write_raw_sheet(
                workbook,
                used_names,
                sheet.get("name", "数据"),
                sheet.get("rows", []),
            )
    else:
        modules = dataset.get("modules", {})
        if modules.get("bp"):
            _write_bp_module(workbook, used_names, dataset, modules["bp"])
        if modules.get("players"):
            _write_player_module(workbook, used_names, modules["players"])
        if modules.get("cooccurrence"):
            _write_cooccurrence_module(
                workbook, used_names, modules["cooccurrence"]
            )
        if modules.get("wards"):
            _write_ward_module(
                workbook,
                used_names,
                dataset.get("team", {}).get("name") or "目标战队",
                modules["wards"],
            )

        team = dataset.get("team", {})
        statistics = dataset.get("statistics", {})
        _write_raw_sheet(
            workbook,
            used_names,
            "原始-战队概况",
            [
                {"字段": display_label(key), "值": display_value(value)}
                for key, value in {**team, **statistics}.items()
            ],
        )
        _write_raw_sheet(
            workbook,
            used_names,
            "原始-当前选手",
            dataset.get("active_players", []),
        )
        _write_raw_sheet(
            workbook,
            used_names,
            "原始-比赛数据",
            dataset.get("matches", []),
        )
        quality = _quality_rows(dataset)
        if quality:
            _write_raw_sheet(workbook, used_names, "数据质量", quality)

    if not workbook.sheetnames:
        _write_raw_sheet(workbook, used_names, "数据", [])
    workbook.save(target)
    return target


def _safe_file_part(value: Any) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', "_", str(value or "目标战队")).strip()
    return text[:60] or "目标战队"


def export_module_workbooks(
    dataset: Dict[str, Any],
    output_dir: str | Path,
    module_keys: Sequence[str] | None = None,
) -> List[Path]:
    """按原脚本模块拆分工作簿，避免不同脚本的 Sheet 混在一起。"""
    target_dir = Path(output_dir).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    modules = dataset.get("modules", {})
    selected = set(module_keys or modules.keys())
    team_name = dataset.get("team", {}).get("name") or "目标战队"
    safe_team = _safe_file_part(team_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outputs: List[Path] = []
    for key, filename in (
        ("bp", f"{safe_team}_BP分析_{timestamp}.xlsx"),
        ("players", f"{safe_team}_选手深度档案_{timestamp}.xlsx"),
        ("cooccurrence", f"{safe_team}_阵容共现分析_{timestamp}.xlsx"),
        ("wards", f"{safe_team}_眼位分析_v1.6b_{timestamp}.xlsx"),
    ):
        module = modules.get(key)
        if key not in selected or not module:
            continue
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_names: set[str] = set()
        if key == "bp":
            _write_bp_module(workbook, used_names, dataset, module)
        elif key == "players":
            _write_player_module(workbook, used_names, module)
        elif key == "cooccurrence":
            _write_cooccurrence_module(workbook, used_names, module)
        else:
            _write_ward_module(workbook, used_names, team_name, module)
        path = target_dir / filename
        workbook.save(path)
        outputs.append(path)
    return outputs


def export_raw_data_workbook(
    dataset: Dict[str, Any], output_path: str | Path
) -> Path:
    """只导出抓取到的基础原始数据，不混入分析脚本的样式化 Sheet。"""
    target = Path(output_path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()
    team = dataset.get("team", {})
    statistics = dataset.get("statistics", {})
    _write_raw_sheet(
        workbook,
        used_names,
        "原始-战队概况",
        [
            {"字段": display_label(key), "值": display_value(value)}
            for key, value in {**team, **statistics}.items()
        ],
    )
    _write_raw_sheet(
        workbook,
        used_names,
        "原始-当前选手",
        dataset.get("active_players", []),
    )
    _write_raw_sheet(
        workbook,
        used_names,
        "原始-比赛数据",
        dataset.get("matches", []),
    )
    quality = _quality_rows(dataset)
    if quality:
        _write_raw_sheet(workbook, used_names, "数据质量", quality)
    if not workbook.sheetnames:
        _write_raw_sheet(workbook, used_names, "原始数据", [])
    workbook.save(target)
    return target

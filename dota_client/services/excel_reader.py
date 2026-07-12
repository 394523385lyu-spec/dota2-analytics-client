from __future__ import annotations

import math
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return None if math.isnan(value) or math.isinf(value) else round(value, 6)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _headers(values: Iterable[Any]) -> List[str]:
    result = []
    used: Counter[str] = Counter()
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value not in (None, "") else f"列{index}"
        used[base] += 1
        result.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return result


def workbook_to_dataset(path: str, max_rows_per_sheet: int = 500) -> Dict[str, Any]:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            first_row = next(iterator)
        except StopIteration:
            sheets.append({"name": worksheet.title, "columns": [], "rows": []})
            continue
        columns = _headers(first_row)
        rows = []
        numeric: Dict[str, List[float]] = {column: [] for column in columns}
        categorical: Dict[str, Counter[str]] = {
            column: Counter() for column in columns
        }
        for row_index, row in enumerate(iterator):
            if row_index >= max_rows_per_sheet:
                break
            record = {
                column: _json_value(row[i] if i < len(row) else None)
                for i, column in enumerate(columns)
            }
            if not any(value not in (None, "") for value in record.values()):
                continue
            rows.append(record)
            for column, value in record.items():
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    numeric[column].append(float(value))
                elif value not in (None, ""):
                    categorical[column][str(value)] += 1
        stats = {}
        for column in columns:
            values = numeric[column]
            if values:
                stats[column] = {
                    "type": "numeric",
                    "count": len(values),
                    "min": round(min(values), 4),
                    "max": round(max(values), 4),
                    "average": round(sum(values) / len(values), 4),
                }
            elif categorical[column]:
                stats[column] = {
                    "type": "categorical",
                    "count": sum(categorical[column].values()),
                    "top_values": categorical[column].most_common(8),
                }
        sheets.append(
            {
                "name": worksheet.title,
                "columns": columns,
                "row_count_loaded": len(rows),
                "worksheet_max_row": worksheet.max_row,
                "statistics": stats,
                "rows": rows,
            }
        )
    workbook.close()
    return {
        "source_file": source.name,
        "source_path": str(source),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }
